import math
import numpy as np
import openpyxl as xl
from openpyxl import Workbook
import cv2
import time
import os
import matplotlib.pyplot as plt
# 另一个可视化轮子
# import chart_studio
import plotly.offline as ptly
import plotly.graph_objs as go

# 打开摄像头，获取模板匹配的结果，得到frame.jpg (要准备模板csmoban.jpg/png)
def video_show(video,ci):
    i = 1
    while True:
        ret1,frame = video.read()
        if not ret1:
            print("视频获取失败！")
            break
        framegray = cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
        # 使用opencv读取图像，直接返回numpy.ndarray 对象，通道顺序为BGR
        # 模板格式可以为jpg和png
        template = cv2.imread('csmoban.jpg')
        # template = cv2.imread('csmoban.png')
        shape = template.shape
        theight , twidth = shape[0] , shape[1]
        # 模板匹配
        result = cv2.matchTemplate(frame,template,cv2.TM_SQDIFF_NORMED)
        cv2.normalize(result,result,0,1,cv2.NORM_MINMAX,-1)
        min_val,max_val,min_loc,max_loc = cv2.minMaxLoc(result)
        # # min_loc：矩形定点 (min_loc[0]+twidth,min_loc[1]+theight)：矩形的宽高 (0,0,225)：矩形的边框颜色；2：矩形边框宽度
        # 画识别框
        cv2.rectangle(frame ,min_loc ,(min_loc[0] + twidth ,min_loc[1] + theight) ,(255 ,0 ,0) ,2)
        # 展示视频数据
        cv2.imshow("Video_show",frame)
        choose_data = framegray[min_loc[1]: (min_loc[1] + theight),min_loc[0]:(min_loc[0] + twidth )]
        # choose_data = frame[min_loc[1]: (min_loc[1] + theight),min_loc[0]:(min_loc[0] + twidth )]
        # 获得模板
        cv2.imshow("result",choose_data)
        if i%ci == 0 :
            cv2.imwrite(f'frame.jpg',choose_data)
            print(f'照片已保存--frame.jpg--：')
            break
        i += 1
        if cv2.waitKey(1) & 0xff == ord("e"):
            break
    video.release()
    cv2.destroyAllWindows()

# 获取指针角度值
def get_pointer_rad(img):
    shape = img.shape
    # image.shape[0], 图片垂直尺寸 // image.shape[1], 图片水平尺寸 // image.shape[2], 图片通道数
    # 图片中心点的坐标 c_y --垂直 c_x --水平
    c_y, c_x, depth = int(shape[0] / 2), int(shape[1] / 2), shape[2]
    # 指针的长度
    l = 1.5*c_x
    src = img.copy()
    list = []
    # 指针画圆形，通过在
    for i in range(361):        # 算法
        x = l * math.cos(i * math.pi / 180) + c_x
        y = l * math.sin(i * math.pi / 180) + c_y
        temp = src.copy()   # 备份
        cv2.line(temp, (c_x, c_y), (int(x), int(y)), (0, 255, 0), thickness=1)  # 在temp上画绿线
        # temp[:,:,1] 就是temp的G通道分量 temp[:,:,:1] == 255 ，为布尔型数据
        c = img[temp[:, :, 1] == 255]
        p = c[c == 0]
        cv2.imshow('temp',temp)
        # point的长度代表匹配程度
        list.append((len(p), i))
        # 如果要求固定检测时间不要太快，可以在这里调慢
        cv2.waitKey(1)
    cv2.destroyAllWindows()
    # 返回一维坐标(points的长度)为最大值时的坐标位置 即为(len,角度)
    # print(list)
    return max(list, key=lambda x: x[0])

# 将图片进行不同的阈值处理,以获得最好的阈值
def getthr(imgc):
    # 每次以不同的阈值来进行图片阈值化
    thres = np.random.randint(40,100)   # 随机数范围
    print(thres)
    imgfan = cv2.threshold(imgc, thres, 255, cv2.THRESH_BINARY)[1]
    cv2.imshow("img",imgfan)
    max = get_pointer_rad(imgfan)
    thr = max[1]
    return thr
# 如果h 不等于1 ，则根据不同的阈值处理结果，得到平均值
# 如果h 等于1，则直接以固定阈值（80）处理，先用平均来测试哪个阈值最好，再用处理的最好的那一个
def get_averg(imgc,h):
    tol = 0
    h = int(h)          # 统计次数
    if h != 1:
        for i in range(h):
            thr = getthr(imgc)
            tol = tol + thr
            print(i+1,end='\n')
        averg = tol / h
    else :
        imgfan = cv2.threshold(imgc,80,255,cv2.THRESH_BINARY)[1]
        max = get_pointer_rad(imgfan)
        thr = max[1]
        averg = thr
    return averg

# 在excel文件中写入数据
def save_xlsx(thr,cdu,timed):
    filex = 'ssdata.xlsx'
    if os.path.exists(filex):
        print(f'--在ssdata.xlsx中写入数据--')
        wb = xl.load_workbook('ssdata.xlsx')
        # 文件表单定位
        sheet = wb['Sheet']
    else:
        wb = Workbook()
        print('--新建一个ssdata.xlsx--')
        sheet = wb['Sheet']

    ws = wb.active
    ws['A1'] = '指针角度'
    ws['B1'] = '温度'
    ws['C1'] = '测试时间'

    maxrow = sheet.max_row + 1
    ws.cell(row=maxrow,column=1,value=thr)
    ws.cell(row=maxrow,column=2,value=cdu)
    ws.cell(row=maxrow,column=3,value=timed)

    wb.save('ssdata.xlsx')

# 重置excel文件
def del_xlsx():
    filex = 'ssdata.xlsx'
    if os.path.exists(filex):
        os.remove(filex)
        wb = Workbook()
        print('--重新生成一个ssdata.xlsx--')
        sheet = wb['Sheet']
    else :
        wb = Workbook()
        print("--新建一个ssdata.xlsx--")

    ws = wb.active
    ws['A1'] = '指针角度'
    ws['B1'] = '温度'
    ws['C1'] = '测试时间'
    wb.save('ssdata.xlsx')

# 读取excel中的数据
def get_cts():
    filex = 'ssdata.xlsx'
    if os.path.exists(filex):
        print(f'--在ssdata.xlsx中写入数据--')
        wb = xl.load_workbook('ssdata.xlsx')
        # 文件表单定位
        sheet = wb['Sheet']
    else:
        wb = Workbook()
        print('--新建一个ssdata.xlsx--')
        sheet = wb['Sheet']
    ws = wb.active
    cdu = []
    for row in ws.iter_rows(min_row=2, min_col=2,max_col=2,max_row=sheet.max_row ):
        for cell in row:
            cdu.append(str(cell.value))
    timed = []
    for row in ws.iter_rows(min_row=2, min_col=3,max_col=3,max_row=sheet.max_row ):
        for cell in row:
            timed.append(cell.value)
    wb.save('ssdata.xlsx')
    return cdu,timed

# 根据读取的以前excel数据生成可视化图形,并保存新的数据
def get_shishi(pingjun):
    # 读取excle中数据
    cts = get_cts()
    cdus = cts[0]
    timeds = cts[1]
    cdu1s = []
    row = True
    while row:
        pd = 1
        imgh = cv2.imread('frame.jpg')
        thr = get_averg(imgh,pingjun)
        print(f'识别的角度值为{thr}')
        # 实物指针度数转化为温度度数
        if thr > 0 and thr <= 45:
            cdu = (4/9) * thr + 100
        elif thr >= 135 and thr < 360:
            cdu = (4/9) * thr - 60
        else :
            cdu = 'error'
            pd = 0

        print(f'转化为温度值{cdu}')
        cdus.append(str(cdu))
        timed = time.strftime("%H:%M:%S", time.localtime())
        timeds.append(timed)

        # 保存数据
        save_xlsx(thr,cdu,timed)

        # plotly 数据可视化
        data = []
        if pd == 1 :
            cdu1 = round(cdu ,1)
        else:
            cdu1 = 'error'
        cdu1s.append(cdu1)
        print(str(cdu1) + "度")
        namee = str(cdu1) + '°C'
        trace1 = go.Scatter(x=timeds,
                            y=cdu1s,
                            mode='lines+markers',  # mode可选'markers','lines','lines+markers'
                            name= namee,
                            marker=dict(size=10,  # 若设为变量则可用散点大小表示变量大小
                                        color='rgba(152, 0, 0, .8)',
                                        line=dict(width=2,
                                                  color='rgb(0, 0, 0)'
                                                  ),
                                        opacity=[]
                                        )
                            )
        data.append(trace1)
        axis_template = dict(
            showgrid=True,  # 网格
            zeroline=True,  # 是否显示基线,即沿着(0,0)画出x轴和y轴
            nticks=20,
            showline=True,
            title='Time',
            mirror='all',
            zerolinecolor="#FF0000"
        )
        ayis_template = dict(
            showgrid=True,  # 网格
            zeroline=True,  # 是否显示基线,即沿着(0,0)画出x轴和y轴
            nticks=20,
            showline=True,
            title='Temp',
            mirror='all',
            zerolinecolor="#FF0000"
        )
        layout = go.Layout(font=dict(family='Courier New, monospace', size=18, color='#3D3D3D'),
                           title='温度值' ,xaxis=axis_template,yaxis=ayis_template
                           )
        fig = go.Figure(data=data, layout=layout)
        ptly.plot(fig, filename='ssdata.html')
        time.sleep(3)
        continue

        # # plot数据可视化
        # # plt.figure(1)
        # plt.clf()  # 清空画布上的所有内容
        # fig1 = plt.figure(num='温度-时间', figsize=(15, 10), dpi=75, facecolor='#FFFFFF', edgecolor='#0000FF')
        # plt.xlabel('Time')
        # plt.ylabel('Temp')
        # # plt.plot(timeds, cdus, ls='-',lw=2,label = "temtim")
        # plt.plot(timeds,cdus,"r-s")
        # # plt.draw()  # 注意此函数需要调用
        # # time.sleep(1)
        # plt.pause(0.01)
        # # plt.show()
        #
        # if len(timeds) >= 18:
        #     timeds = []
        #     cdus = []
        #     plt.clf()
        # else :
        #     continue


