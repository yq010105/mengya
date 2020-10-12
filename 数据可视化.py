import openpyxl as xl
import os
from openpyxl import Workbook
import plotly.offline as ptly
import plotly.graph_objs as go

def get_cts():
    filex = 'ssdata.xlsx'
    if os.path.exists(filex):
        print(f'--在ssdata.xlsx中写入数据--')
        wb = xl.load_workbook('ssdata.xlsx')
        # 文件表单定位
        sheet = wb['Sheet']
    else:
        wb = Workbook()
        print('--新建一个ssdata.xlsx--')
        sheet = wb['Sheet']
    ws = wb.active
    cdu = []
    for row in ws.iter_rows(min_row=2, min_col=2,max_col=2,max_row=sheet.max_row ):
        for cell in row:
            cdu.append(str(cell.value))
    timed = []
    for row in ws.iter_rows(min_row=2, min_col=3,max_col=3,max_row=sheet.max_row ):
        for cell in row:
            timed.append(cell.value)
    wb.save('ssdata.xlsx')
    return cdu,timed

cdutimed= get_cts()
cdu1s = cdutimed[0]
timeds = cdutimed[1]
data = []

trace1 = go.Scatter(x=timeds,
                    y=cdu1s,
                    mode='lines+markers',  # mode可选'markers','lines','lines+markers'
                    marker=dict(size=10,  # 若设为变量则可用散点大小表示变量大小
                                color='rgba(152, 0, 0, .8)',
                                line=dict(width=2,
                                            color='rgb(0, 0, 0)'
                                            ),
                                opacity=[]
                                )
                    )
data.append(trace1)
axis_template = dict(
    showgrid=True,  # 网格
    zeroline=True,  # 是否显示基线,即沿着(0,0)画出x轴和y轴
    nticks=20,
    showline=True,
    title='Time',
    mirror='all',
    zerolinecolor="#FF0000"
)
ayis_template = dict(
    showgrid=True,  # 网格
    zeroline=True,  # 是否显示基线,即沿着(0,0)画出x轴和y轴
    nticks=20,
    showline=True,
    title='Temp',
    mirror='all',
    zerolinecolor="#FF0000"
)
layout = go.Layout(font=dict(family='Courier New, monospace', size=18, color='#3D3D3D'),
                    title='温度值' ,xaxis=axis_template,yaxis=ayis_template
                    )
fig = go.Figure(data=data, layout=layout)
ptly.plot(fig, filename='ssdata.html')